from selenium import webdriver
from TestSuiteReporter import TestSuiteReporter
import logging
from typing import Callable, Union
import mysql.connector
from mysql.connector import Error
import random, string, time
from Outlook import Outlook_App
from openpyxl import load_workbook

def report_event_and_log(
    driver, 
    message: str, 
):
    logging.getLogger(driver.loggingID).info(message)
    driver.reporter[driver.testID].reportEvent(message, False, "")

def load_excel_sheet(driver, rowName):
    wb = load_workbook(filename = 'AF_Register_Bank_Member.xlsx', data_only=True)
    sheet = wb['RegisterBank']
    driver.data = {}
    for row in sheet.rows:
        if row[0].value == rowName:
            for i in range (2, 22):
                driver.data[sheet.cell(column=i, row=1).value] = sheet.cell(column=i, row=row[0].row).value
            logging.getLogger(driver.loggingID).info("Loaded excel data")

def check_for_responsive(driver):
    logging.getLogger(driver.loggingID).info("Checking for responsive webpage")
    if driver.get_window_size()['width'] < 550:
        driver.responsive = True
        logging.getLogger(driver.loggingID).info("Detected responsive")
    else:
        driver.responsive = False
        logging.getLogger(driver.loggingID).info("Didn't detect responsive")

def check_outlook_confirmation(driver, timeout=60):
    timer = 0
    obj = Outlook_App()
    while obj.search_by_subject("Welcome", 6) == -1 and timer < timeout:
        timer += 1
        time.sleep(1)
        logging.getLogger(driver.loggingID).info(
            "Waiting for Outlook confirmation " + str(timer) + "/" + str(timeout) + "s"
        )

    if obj.search_by_subject("Welcome", 6) == -1:
        driver.reporter[driver.testID].reportStep(
            "Check if confirmation email was received",
            "Confirmation email was received",
            "Confirmation email wasn't received after " + str(timeout) + " seconds",
            False
        )
        logging.getLogger(driver.loggingID).info("Confirmation email wasn't received after " + str(timeout) + " seconds")

    else:
        driver.reporter[driver.testID].reportStep(
            "Check if confirmation email was received",
            "Confirmation email was received",
            "Confirmation email was received after " + str(timer) + " seconds",
            True
        )
        logging.getLogger(driver.loggingID).info("Confirmation email was received after " + str(timer) + " seconds")




#SQL Utilites
def create_aline_sql_connection(driver):
    connection = None
    try:
        connection = mysql.connector.connect(
            host='uftcapstone-db.c1ddjzxizuua.us-east-1.rds.amazonaws.com',
            user='uftcapstone',
            passwd='!A&8vYOKSUO&X9Zt',
            database='alinedb'
        )
        logging.getLogger(driver.loggingID).info("MySQL Database connection sucessful")
    except Error as err:
        logging.getLogger(driver.loggingID).info(f"Error: '{err}'")

    return connection

def read_query(driver, connection, query):
    cursor = connection.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        return result
    except Error as err:
        logging.getLogger(driver.loggingID).info(f"Error: '{err}'")

def execute_query(driver, connection, query):
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        connection.commit()
        logging.getLogger(driver.loggingID).info("Query successful")
    except Error as err:
        logging.getLogger(driver.loggingID).info(f"Error '{err}'")

def find_and_update_email(driver, email):
    connection = create_aline_sql_connection(driver)
    query = "SELECT * FROM applicant WHERE email='" + email + "'"
    results = read_query(driver, connection, query)
    if len(results) > 0:
        random_email = ''.join(random.choices(string.ascii_lowercase, k=12)) + '@' + ''.join(random.choices(string.ascii_lowercase, k=12)) + '.com'
        update_str = "UPDATE applicant SET email='" + random_email + "' WHERE email='" + email + "'"
        execute_query(driver, connection, update_str)
