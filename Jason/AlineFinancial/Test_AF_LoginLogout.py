import unittest
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
#from selenium.webdriver.chrome.options import Options
from selenium.webdriver.firefox.options import Options
from AF_AdminLogin import AF_Login
import HTML_Reporting
import time
import random
import string
import openpyxl
import boto3
from S3_Tool import upload_file
import os
from os.path import basename
import zipfile
##import win32com.client as win32


class Test_LoginLogout(unittest.TestCase):

    def __init__(self, *args, **kwargs):
        super(Test_LoginLogout, self).__init__(*args, **kwargs)
##        self.timestr = time.strftime("%Y-%m-%d--%I_%M_%S%p")
##        #self.reporter = HTML_Reporting.TestSuiteReporter(self.timestr, "D:\\TestingResources\\AlineFinancial\\TestResults\\", "Jason")
##        self.reporter = HTML_Reporting.TestSuiteReporter(self.timestr, "./", "Jason")
##        #self.screenshotPath = "D:\\TestingResources\\AlineFinancial\\TestResults\\.screenshots\\"
##        self.screenshotPath = ""
##        #self.path = "D:\\TestingResources\\AlineFinancial\\DataSheets\\InputData.xlsm"
##        self.path = "InputData.xlsm"
##        self.xl = win32.Dispatch("Excel.Application")
##        #self.xl.Interactive = False
##        #self.xl.Visible = False
##        xlbook = self.xl.Workbooks.Open(self.path)
##        self.xl.Application.Run("InputData.xlsm!Module1.ResetLastCell()")
##        xlbook.Save()
##        self.xl.Application.Quit()

    @classmethod
    def setUpClass(cls):
        cls.imageFolders = []
        cls.timestr = time.strftime("%Y-%m-%d--%I_%M_%S%p")
        cls.reporter = HTML_Reporting.TestSuiteReporter(cls.timestr, "./", "Jason")
        cls.screenshotPath = ""
        cls.path = "InputData.xlsm"


    def setUp(self): 
        options = Options()
        options.headless = True
        #self.driver = webdriver.Chrome(options=options)
        self.driver = webdriver.Firefox(options=options)
        #self.imageFolders = []

    def test_001_login(self):
        #setup
        path = self.path
        wb = openpyxl.load_workbook(path)
        ws = wb["TC_001"]
        reporter = self.reporter
        driver = self.driver
        #repeat test with each data row
        for row in ws.iter_rows(min_row=2):

            print("##############################################")

            #set driver size
            res = row[2].value.split("x")
            if len(res)==2:
                driver.set_window_size(int(res[0]), int(res[1]))
##            elif res[0]=="fullscreen":
##                driver.fullscreen_window()
##                driver.manage().window().maximize();
            elif res[0] is None:
                pass

            #print(driver.get_window_size())
            
            #set up reporter
            testID="TC_001" +"_"+ ''.join(random.choices(string.ascii_lowercase, k=5))
            self.imageFolders.append(testID)
            reporter.addTestCase(testID, "JS_TC_001", "Login to Aline Financial - Admin")
            reporter[testID].reportEvent("Set resolution for testing",False,res)
            #login process
            loginObj = AF_Login(driver)
            loginObj.Launch_Login_Page()
            loginObj.AF_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)
            loginObj.AF_logout(reporter[testID], self.screenshotPath)


    def test_002_login_neg(self):
        #setup
        path = self.path
        wb = openpyxl.load_workbook(path)
        ws = wb["TC_002"]
        reporter = self.reporter
        driver = self.driver
        #repeat test with each data row
        for row in ws.iter_rows(min_row=2):

            print("##############################################")
            #set driver size
            res = row[2].value.split("x")
            if len(res)==2:
                driver.set_window_size(int(res[0]), int(res[1]))
##            elif res[0]=="fullscreen":
##                driver.fullscreen_window()
##                driver.manage().window().maximize();
            elif res[0] is None:
                pass
            
            #set up reporter
            testID="TC_002" +"_"+ ''.join(random.choices(string.ascii_lowercase, k=5))
            self.imageFolders.append(testID)
            reporter.addTestCase(testID, "JS_TC_002", "Attempt to login to Aline Financial - Admin with bad credentials")
            reporter[testID].reportEvent("Set resolution for testing",False,res)
            #login process
            loginObj = AF_Login(driver)
            loginObj.Launch_Login_Page()
            loginObj.AF_bad_login(reporter[testID], self.screenshotPath, row[0].value, row[1].value)



    def tearDown(self):
        #self.driver.close()
        self.driver.quit()


    @classmethod
    def tearDownClass(cls):
        del cls.reporter
        zipObj = zipfile.ZipFile(cls.timestr+".zip", 'w')
        zipObj.write(cls.timestr + ".html")
        for folder in cls.imageFolders:
            for image in os.listdir("./.screenshots/"+folder+"/"):
                zipObj.write("./.screenshots/"+folder+"/"+image)
        zipObj.close()
        upload_file(cls.timestr+".zip","scrumbags-reports")

if __name__ == "__main__":

    unittest.main(warnings='ignore')

